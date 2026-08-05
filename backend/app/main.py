import os
import io
import time
import json
import asyncio
import logging
import urllib.request
import subprocess
from typing import Optional
from dotenv import load_dotenv

# Automatically load environment variables from .env file
load_dotenv()
load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, Response, FileResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.errors import IgnoredError, IgnoredErrors

from . import ocr_engine
from .pii_masker import PIIMasker
from . import llm_connector
from . import db_manager
from .parsers.invoice_parser import clean_final_invoice_data

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("backend_main")

app = FastAPI(
    title="Smart Document Organizer IDP Backend",
    description="Decoupled FastAPI IDP Backend for Document Scanning, PII Masking, LLM & Excel Export",
    version="2.0.0"
)

# Enable CORS for Next.js and Streamlit frontends
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
def startup_event():
    db_manager.init_db()
    # Auto-start Ollama Local AI Vision server if not already running
    try:
        ollama_env = os.environ.copy()
        if "OLLAMA_MODELS" not in ollama_env:
            ollama_env["OLLAMA_MODELS"] = r"D:\My Data\Project\OllamaModels"
        
        try:
            urllib.request.urlopen("http://localhost:11434/api/tags", timeout=1)
            logger.info("🧠 Local Ollama AI Vision server is already ONLINE at http://localhost:11434")
        except Exception:
            ollama_bin = os.path.expanduser('~') + r"\AppData\Local\Programs\Ollama\ollama.exe"
            alt_bin = r"D:\My Data\Project\Ollama\ollama.exe"
            target_bin = ollama_bin if os.path.exists(ollama_bin) else (alt_bin if os.path.exists(alt_bin) else None)
            if target_bin:
                logger.info("🚀 Auto-launching Ollama Local AI Vision server in background...")
                subprocess.Popen([target_bin, "serve"], env=ollama_env)
    except Exception as e:
        logger.warning(f"Ollama auto-start check error: {e}")

@app.get("/api/health")
def health_check():
    return {"status": "online", "system": "Smart Document Organizer IDP Backend"}

@app.post("/api/idp/process-stream")
async def process_document_stream(
    file: UploadFile = File(...),
    doc_type: str = Form("auto"),
    is_confidential: bool = Form(True),
    llm_engine: str = Form("local"),
    custom_api_key: str = Form("")
):
    """
    Asynchronous Server-Sent Events (SSE) streaming endpoint.
    Streams progress step-by-step with AI Auto-Classification.
    """
    file_bytes = await file.read()
    file_name = file.filename or "uploaded_doc.jpg"

    async def event_generator():
        start_t = time.time()
        try:
            # Step 0: Upload
            yield f"data: {json.dumps({'step': 0, 'status': 'Upload Selesai', 'message': f'File {file_name} berhasil diunggah'})}\n\n"
            await asyncio.sleep(0.3)

            is_local_vision = llm_engine in ["ollama", "local_vision", "moondream", "qwen2-vl"]
            is_direct_vision = (llm_engine in ["cloud_vision", "direct_vision", "vision_ai", "gemini", "openai", "claude"] or is_local_vision) and (not is_confidential or is_local_vision)
            has_cloud_key = is_local_vision or bool(custom_api_key or os.getenv("GEMINI_API_KEY") or os.getenv("OPENAI_API_KEY") or os.getenv("ANTHROPIC_API_KEY"))

            if is_direct_vision:
                # Option B: Direct Vision AI Mode (Cloud or 100% Local Ollama)
                prov_name = "Ollama Local Vision (Qwen2.5-VL)" if is_local_vision else (llm_engine.upper() if llm_engine not in ["cloud_vision", "direct_vision"] else "Cloud Vision AI")
                if has_cloud_key:
                    yield f"data: {json.dumps({'step': 1, 'status': f'Vision AI Engine ({prov_name})', 'message': f'Mengirimkan gambar Base64 ke Vision Model {prov_name}...'})}\n\n"
                else:
                    yield f"data: {json.dumps({'step': 1, 'status': 'Direct Vision (Fallback Mode)', 'message': '⚠️ Cloud API Key belum diset di .env. Menggunakan High-Res Preprocessing Engine...'})}\n\n"
                
                extracted_json, effective_doc_type = llm_connector.analyze_document_image(file_bytes, doc_type, provider=llm_engine, custom_api_key=custom_api_key, raw_text="")
                raw_ocr = f"[Direct Vision AI Reading ({prov_name}) - Raw Image Base64 Processed]"
                masked_text = raw_ocr
                mask_map = {}
                
                # Format & Harmonize Vision AI Output Currency & Date Standards
                if effective_doc_type in ["invoice", "vendor_doc", "vendor", "po"]:
                    final_data = clean_final_invoice_data(extracted_json, raw_ocr)
                else:
                    final_data = extracted_json
                await asyncio.sleep(0.3)
            else:
                # Option A: Privacy & Confidential Mode (100% Secure Local OCR + PII Sensor Vault)
                yield f"data: {json.dumps({'step': 1, 'status': 'Proses OCR & Klasifikasi AI', 'message': 'Menjalankan EasyOCR & Auto-Classifier AI...'})}\n\n"
                ocr_res = ocr_engine.process_document(file_bytes, file_name)
                raw_ocr = ocr_res.get('text', '')

                effective_doc_type = doc_type.lower().strip()
                if effective_doc_type in ["auto", "", "none"]:
                    effective_doc_type = llm_connector.detect_document_type(raw_ocr)
                    logger.info(f"Auto-classified raw_ocr -> {effective_doc_type}")
                await asyncio.sleep(0.3)

                # Step 2: PII Masking
                masker = PIIMasker()
                yield f"data: {json.dumps({'step': 2, 'status': 'PII Masking', 'message': f'Memproses sensor PII untuk kategori {effective_doc_type.upper()}...'})}\n\n"
                if is_confidential:
                    masked_text, mask_map = masker.mask(raw_ocr)
                else:
                    masked_text = raw_ocr
                    mask_map = {}
                await asyncio.sleep(0.3)

                # Step 3: LLM Processing
                yield f"data: {json.dumps({'step': 3, 'status': 'Proses LLM AI', 'message': f'Menganalisis entitas {effective_doc_type.upper()}...'})}\n\n"
                extracted_json, _ = llm_connector.analyze_document_text(masked_text, effective_doc_type)
                await asyncio.sleep(0.3)

                # Step 4: Validasi & Unmasking
                yield f"data: {json.dumps({'step': 4, 'status': 'Validasi & Unmask', 'message': 'Melakukan dekripsi PII Vault...'})}\n\n"
                if is_confidential and mask_map:
                    final_data = masker.unmask(extracted_json)
                else:
                    final_data = extracted_json

                if effective_doc_type in ["invoice", "vendor_doc", "vendor", "po"]:
                    final_data = clean_final_invoice_data(final_data, raw_ocr)
                await asyncio.sleep(0.3)

            # Step 5: Format Excel
            yield f"data: {json.dumps({'step': 5, 'status': 'Format Excel', 'message': 'Menyiapkan struktur sheet Excel...'})}\n\n"
            await asyncio.sleep(0.2)

            process_time = round(time.time() - start_t, 2)

            # Step 6: Final Complete Result
            result_payload = {
                "file_name": file_name,
                "doc_type": effective_doc_type,
                "is_auto_detected": (doc_type == "auto"),
                "is_confidential": is_confidential,
                "llm_engine": llm_engine,
                "raw_ocr": raw_ocr,
                "masked_text": masked_text,
                "mask_map": mask_map,
                "llm_json": extracted_json,
                "final_data": final_data,
                "process_time_seconds": process_time
            }
            yield f"data: {json.dumps({'step': 6, 'status': 'Hasil Selesai', 'data': result_payload})}\n\n"

        except Exception as e:
            logger.error(f"Error in stream processing: {e}")
            yield f"data: {json.dumps({'step': -1, 'status': 'Error', 'error': str(e)})}\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")

class APIKeysPayload(BaseModel):
    gemini_key: Optional[str] = None
    openai_key: Optional[str] = None
    anthropic_key: Optional[str] = None

@app.get("/api/settings/keys")
def get_api_keys():
    """Returns masked versions of stored API keys."""
    def mask_k(val):
        if not val or "isi_dengan" in val:
            return ""
        return val[:6] + "..." + val[-4:] if len(val) > 10 else "******"

    return {
        "gemini_key": mask_k(os.getenv("GEMINI_API_KEY", "")),
        "openai_key": mask_k(os.getenv("OPENAI_API_KEY", "")),
        "anthropic_key": mask_k(os.getenv("ANTHROPIC_API_KEY", "")),
        "has_gemini": bool(os.getenv("GEMINI_API_KEY") and "isi_dengan" not in os.getenv("GEMINI_API_KEY")),
        "has_openai": bool(os.getenv("OPENAI_API_KEY") and "isi_dengan" not in os.getenv("OPENAI_API_KEY")),
        "has_anthropic": bool(os.getenv("ANTHROPIC_API_KEY") and "isi_dengan" not in os.getenv("ANTHROPIC_API_KEY")),
    }

@app.post("/api/settings/keys")
def update_api_keys(payload: APIKeysPayload):
    """Updates API keys in memory and writes to backend/.env file."""
    env_path = os.path.join(os.path.dirname(__file__), "..", ".env")
    env_data = {}
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                if "=" in line and not line.strip().startswith("#"):
                    k, v = line.strip().split("=", 1)
                    env_data[k.strip()] = v.strip()

    if payload.gemini_key is not None and payload.gemini_key and not payload.gemini_key.startswith("***"):
        os.environ["GEMINI_API_KEY"] = payload.gemini_key
        env_data["GEMINI_API_KEY"] = payload.gemini_key

    if payload.openai_key is not None and payload.openai_key and not payload.openai_key.startswith("***"):
        os.environ["OPENAI_API_KEY"] = payload.openai_key
        env_data["OPENAI_API_KEY"] = payload.openai_key

    if payload.anthropic_key is not None and payload.anthropic_key and not payload.anthropic_key.startswith("***"):
        os.environ["ANTHROPIC_API_KEY"] = payload.anthropic_key
        env_data["ANTHROPIC_API_KEY"] = payload.anthropic_key

    with open(env_path, "w", encoding="utf-8") as f:
        f.write("# Smart Document Organizer - Environment Configurations & Cloud API Keys\n")
        for k, v in env_data.items():
            f.write(f"{k}={v}\n")

    return {"status": "success", "message": "API Keys updated successfully!"}

@app.get("/api/documents")
def list_documents(search: Optional[str] = Query(None), doc_type: Optional[str] = Query(None)):
    """API endpoint to get saved documents from SQLite."""
    return db_manager.get_documents(search_query=search, doc_type=doc_type)

@app.post("/api/documents/save")
async def save_document(
    file: UploadFile = File(...),
    doc_type: str = Form("ktp"),
    raw_text: str = Form(""),
    masked_text: str = Form(""),
    json_data: str = Form("{}"),
    is_confidential: bool = Form(True),
    llm_engine: str = Form("Local Rule Parser"),
    process_time_seconds: float = Form(0.0)
):
    """API endpoint to save processed document into SQLite DB and disk."""
    file_bytes = await file.read()
    try:
        parsed_json = json.loads(json_data)
    except Exception:
        parsed_json = {"raw": json_data}

    saved_doc_type = doc_type
    if saved_doc_type.lower() in ["auto", "", "none"] and isinstance(parsed_json, dict):
        po_val = str(parsed_json.get("po_number", "")).strip()
        doc_cat = str(parsed_json.get("document_category", "") or parsed_json.get("document_type", "")).lower()
        if (po_val and po_val.upper() not in ["N/A", "NONE", "NULL", "", "-"]) or "vendor" in doc_cat:
            saved_doc_type = "vendor"
        elif "invoice_number" in parsed_json or "vendor_name" in parsed_json or "customer_name" in parsed_json:
            saved_doc_type = "invoice"
        elif "nik" in parsed_json or ("id_number" in parsed_json and "religion" in parsed_json):
            saved_doc_type = "ktp"
        elif "passport_number" in parsed_json or "nationality" in parsed_json:
            saved_doc_type = "passport"
        elif "contact_name" in parsed_json or "company_name" in parsed_json:
            saved_doc_type = "business_card"
        else:
            saved_doc_type = "general"

    doc_id = db_manager.save_document(
        file_name=file.filename or "document.jpg",
        file_bytes=file_bytes,
        doc_type=saved_doc_type,
        raw_text=raw_text,
        masked_text=masked_text,
        json_data=parsed_json,
        is_confidential=is_confidential,
        llm_engine=llm_engine,
        process_time_seconds=process_time_seconds
    )
    return {"status": "success", "id": doc_id}

@app.delete("/api/documents/clear-all")
def clear_all_documents():
    """API endpoint to delete all documents."""
    db_manager.delete_all_documents()
    return {"status": "success", "message": "All documents deleted."}

@app.delete("/api/documents/{doc_id}")
def delete_document(doc_id: int):
    """API endpoint to delete a single document."""
    success = db_manager.delete_document(doc_id)
    if not success:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"status": "success", "message": f"Document #{doc_id} deleted."}

@app.get("/api/export/excel")
def export_multi_sheet_excel(doc_type: Optional[str] = Query("All")):
    """
    Generates and downloads a multi-sheet formatted Excel workbook (.xlsx).
    """
    docs = db_manager.get_documents(doc_type=doc_type)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    # Sheet configurations
    sheet_configs = {
        "KTP": ["ID Dokumen", "File Dokumen", "NIK", "Nama Lengkap", "Tempat/Tgl Lahir", "Jenis Kelamin", "Alamat", "RT/RW", "Kel/Desa", "Kecamatan", "Agama", "Status Perkawinan", "Pekerjaan", "Waktu Ekstraksi (s)", "Tanggal Diunggah"],
        "Paspor": ["ID Dokumen", "File Dokumen", "Tipe Dokumen", "Tipe Paspor", "Kode Negara", "No Paspor", "Nama Lengkap", "Tempat Lahir", "Tanggal Lahir", "Jenis Kelamin", "Kewarganegaraan", "Tanggal Pengeluaran", "Habis Berlaku", "No. Reg.", "Kantor Pengeluar", "MRZ Code", "Waktu Ekstraksi (s)", "Tanggal Diunggah"],
        "Kartu Nama": ["ID Dokumen", "File Dokumen", "Nama Kontak", "Jabatan", "Perusahaan", "No Telepon", "Email", "Website", "Waktu Ekstraksi (s)", "Tanggal Diunggah"],
        "Invoice & Struk": ["ID Dokumen", "File Dokumen", "Nama Vendor / Toko", "Ditujukan Kepada", "No Invoice", "Tanggal Invoice", "Batas Tempo", "Subtotal", "Pajak (Tax)", "Total Amount", "Mata Uang", "Waktu Ekstraksi (s)", "Tanggal Diunggah"],
        "Dokumen Vendor & PO": ["ID Dokumen", "File Dokumen", "Nama Vendor", "Pemesan / Ditujukan", "No PO / Kontrak", "No Invoice", "No Surat Jalan", "Tgl Pesan", "Tgl Datang", "Mata Uang", "Subtotal", "PPN (%)", "Pajak (Rp/USD)", "Total Amount", "Waktu Ekstraksi (s)", "Tanggal Diunggah"],
        "Dokumen Bisnis & Pajak": ["ID Dokumen", "File Dokumen", "Judul Dokumen", "NPWP", "NIB", "No Sertifikat / Izin", "Tanggal Terbit", "Ringkasan", "Waktu Ekstraksi (s)", "Tanggal Diunggah"]
    }

    category_map = {
        "ktp": "KTP",
        "id_card": "KTP",
        "passport": "Paspor",
        "business_card": "Kartu Nama",
        "invoice": "Invoice & Struk",
        "vendor": "Dokumen Vendor & PO",
        "vendor_doc": "Dokumen Vendor & PO",
        "po": "Dokumen Vendor & PO",
        "general": "Dokumen Bisnis & Pajak"
    }

    categorized_docs = {title: [] for title in sheet_configs.keys()}
    for d in docs:
        cat_title = category_map.get(d['type'].lower(), "Dokumen Bisnis & Pajak")
        categorized_docs[cat_title].append(d)

    for sheet_title, headers in sheet_configs.items():
        ws = wb.create_sheet(title=sheet_title)
        ws.ignored_errors = IgnoredErrors()
        ws.ignored_errors.ignoredError.append(IgnoredError(sqref="A1:Z5000", numberStoredAsText=True))

        ws.append(headers)

        # Style header row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        doc_list = categorized_docs[sheet_title]
        for d in doc_list:
            j = d['json_data']
            dt = d['created_at']
            proc_time = f"{round(float(d.get('process_time_seconds', 0.0) or 0.0), 2)}s"

            if sheet_title in ["Invoice & Struk", "Dokumen Vendor & PO"]:
                j = clean_final_invoice_data(j, d.get('raw_text', ''))

            if sheet_title == "KTP":
                pob = j.get('place_of_birth', 'N/A')
                dob = j.get('date_of_birth', 'N/A')
                pob_dob = f"{pob}, {dob}" if pob != 'N/A' else dob
                row = [d['id'], d['name'], j.get('id_number', 'N/A'), j.get('full_name', 'N/A'), pob_dob, j.get('gender', 'N/A'), j.get('address', 'N/A'), j.get('rt_rw', 'N/A'), j.get('kel_desa', 'N/A'), j.get('kecamatan', 'N/A'), j.get('religion', 'N/A'), j.get('marital_status', 'N/A'), j.get('occupation', 'N/A'), proc_time, dt]
            elif sheet_title == "Paspor":
                pass_num = j.get('passport_number') or j.get('id_number') or 'N/A'
                row = [d['id'], d['name'], j.get('document_type', 'Passport'), j.get('passport_type', 'P'), j.get('country_code', 'IDN'), pass_num, j.get('full_name', 'N/A'), j.get('place_of_birth', 'N/A'), j.get('date_of_birth', 'N/A'), j.get('gender', 'N/A'), j.get('nationality', 'INDONESIA'), j.get('issue_date', 'N/A'), j.get('expiry_date', 'N/A'), j.get('registration_no', 'N/A'), j.get('issuing_office', 'N/A'), j.get('mrz_code', 'N/A'), proc_time, dt]
            elif sheet_title == "Kartu Nama":
                row = [d['id'], d['name'], j.get('contact_name', 'N/A'), j.get('job_title', 'N/A'), j.get('company_name', 'N/A'), j.get('phone_number', 'N/A'), j.get('email_address', 'N/A'), j.get('website_url', 'N/A'), proc_time, dt]
            elif sheet_title == "Invoice & Struk":
                row = [d['id'], d['name'], j.get('vendor_name', 'N/A'), j.get('customer_name', 'N/A'), j.get('invoice_number', 'N/A'), j.get('invoice_date', 'N/A'), j.get('due_date', 'N/A'), j.get('subtotal', '0.00'), j.get('tax', '0.00'), j.get('total_amount', 'N/A'), j.get('currency', 'IDR'), proc_time, dt]
            elif sheet_title == "Dokumen Vendor & PO":
                po_val = j.get('po_number')
                po_str = "N/A" if not po_val or str(po_val).strip() in ["", "null", "None"] else str(po_val).strip()
                tax_pct = str(j.get('tax_percent') or '0%').strip()
                row = [d['id'], d['name'], j.get('vendor_name', 'N/A'), j.get('customer_name', 'N/A'), po_str, j.get('invoice_number', 'N/A'), j.get('delivery_order_number', 'N/A'), j.get('order_date', 'N/A'), j.get('delivery_date', 'N/A'), j.get('currency', 'IDR'), j.get('subtotal', '0.00'), tax_pct, j.get('tax', '0.00'), j.get('total_amount', 'N/A'), proc_time, dt]
            else:
                row = [d['id'], d['name'], j.get('document_title', 'N/A'), j.get('tax_id_npwp', 'N/A'), j.get('business_license_nib', 'N/A'), j.get('certificate_number', 'N/A'), j.get('issue_date', 'N/A'), j.get('summary', 'N/A'), proc_time, dt]

            ws.append(row)
            curr_row_idx = ws.max_row
            for col_i, val_i in enumerate(row, start=1):
                cell = ws.cell(row=curr_row_idx, column=col_i)
                s_val = str(val_i or '').strip()
                if s_val.isdigit() and len(s_val) > 1 and s_val.startswith('0'):
                    cell.value = int(s_val)
                    cell.number_format = "0" * len(s_val)

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 1. Dedicated Items Sheet for Invoice & Struk
    ws_inv_items = wb.create_sheet(title="Item Invoice & Struk")
    ws_inv_items.ignored_errors = IgnoredErrors()
    ws_inv_items.ignored_errors.ignoredError.append(IgnoredError(sqref="A1:Z5000", numberStoredAsText=True))
    inv_item_headers = ["ID Dokumen", "File Dokumen", "Nama Vendor / Toko", "Ditujukan Kepada", "No Invoice", "No Item", "SKU / Kode", "Deskripsi Barang / Item", "Jumlah (Qty)", "Satuan", "Harga Satuan", "Total Harga Item", "Tanggal Diunggah"]
    ws_inv_items.append(inv_item_headers)
    for col_idx in range(1, len(inv_item_headers) + 1):
        cell = ws_inv_items.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 2. Dedicated Items Sheet for Vendor & PO
    ws_ven_items = wb.create_sheet(title="Item Vendor & PO")
    ws_ven_items.ignored_errors = IgnoredErrors()
    ws_ven_items.ignored_errors.ignoredError.append(IgnoredError(sqref="A1:Z5000", numberStoredAsText=True))
    ven_item_headers = ["ID Dokumen", "File Dokumen", "Nama Vendor", "Pemesan (Ditujukan)", "No PO / Kontrak", "No Invoice", "No Item", "SKU / Kode", "Deskripsi Barang / Item", "Jumlah (Qty)", "Satuan", "Harga Satuan", "Total Harga Item", "Tanggal Diunggah"]
    ws_ven_items.append(ven_item_headers)
    for col_idx in range(1, len(ven_item_headers) + 1):
        cell = ws_ven_items.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Populate item breakdown sheets
    for d in docs:
        j = d['json_data']
        dt = d['created_at']
        doc_type_lower = d['type'].lower()
        items = j.get('items', [])
        
        target_ws = ws_ven_items if doc_type_lower in ["vendor", "vendor_doc", "po"] else ws_inv_items
        
        if isinstance(items, list) and len(items) > 0:
            for idx, it in enumerate(items):
                if isinstance(it, dict):
                    if target_ws == ws_inv_items:
                        item_row = [
                            d['id'], d['name'], j.get('vendor_name', 'N/A'), j.get('customer_name', 'N/A'),
                            j.get('invoice_number', 'N/A'), it.get('no', idx + 1), it.get('sku', '-'),
                            it.get('description', it.get('item_name', '-')), it.get('qty', '1'), it.get('unit', 'pcs'),
                            it.get('unit_price', '0'), it.get('total', '0'), dt
                        ]
                    else:
                        item_row = [
                            d['id'], d['name'], j.get('vendor_name', 'N/A'), j.get('customer_name', 'N/A'),
                            j.get('po_number', 'N/A'), j.get('invoice_number', 'N/A'), it.get('no', idx + 1),
                            it.get('sku', '-'), it.get('description', it.get('item_name', '-')), it.get('qty', '1'),
                            it.get('unit', 'pcs'), it.get('unit_price', '0'), it.get('total', '0'), dt
                        ]
                    target_ws.append(item_row)

    for ws_cur in [ws_inv_items, ws_ven_items]:
        for col in ws_cur.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_cur.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        'Content-Disposition': 'attachment; filename="Rekap_Dokumen_Multi_Sheet.xlsx"'
    }
    return Response(output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
